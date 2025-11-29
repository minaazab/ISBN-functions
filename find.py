from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def finds(cell, sheet_num):
    for col in sheet_num.iter_cols(min_col=1, max_col=4):
        for num in col:
            if(num.value == cell):
                return True
    return False


def process_find(file_path):
    wb = load_workbook(file_path)
    sheet = wb['Sheet1']
    sheet2 = wb['Sheet2']
    y = 1
    while sheet['A' + str(y)].value is not None:    
        cell = sheet['A'+ str(y)]
        if(finds(cell.value, sheet2)):
            print(cell.value)
            cell.fill = yellow
        y += 1
    wb.save(file_path)
    return file_path

def open_sesame(file_path):
    os.startfile(file_path)
